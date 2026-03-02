#!/usr/bin/env python3
"""Import customers + suppliers workbook into Azure SQL."""

from __future__ import annotations

import argparse
import datetime as dt
import os
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

SOURCE_SYSTEM = "customers_suppliers_workbook"


@dataclass
class CustomerContactRow:
    source_contact_id: int | None
    name: str
    contact_name: str | None
    organization_name: str | None
    address: str | None
    phone: str | None
    email: str | None
    tax_id: str | None
    contact_type: str | None
    source_channel: str | None
    shipping_address: str | None
    shipping_email: str | None
    shipping_phone: str | None
    notes: str | None


@dataclass
class SupplierContactRow:
    source_contact_id: int | None
    name: str
    contact_name: str | None
    organization_name: str | None
    branch_name: str | None
    address: str | None
    phone: str | None
    email: str | None
    tax_id: str | None
    source_channel: str | None
    shipping_address: str | None
    shipping_email: str | None
    shipping_phone: str | None
    notes: str | None


def normalize_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        text = str(value).strip()
    elif isinstance(value, float):
        if abs(value - int(value)) < 1e-9:
            text = str(int(value))
        else:
            text = str(value).strip()
    else:
        text = str(value).strip()
    if not text or text in {"-", "--", "#"}:
        return None
    return text


def parse_int(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        rounded = int(round(value))
        return rounded if abs(value - rounded) < 1e-9 else None

    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def pick_name(contact_name: str | None, organization_name: str | None, fallback: str) -> str:
    for value in (organization_name, contact_name):
        if value and value.strip():
            return value.strip()[:180]
    return fallback


def parse_customers_sheet(workbook: openpyxl.Workbook) -> list[CustomerContactRow]:
    if "Customer" not in workbook.sheetnames:
        return []

    sheet = workbook["Customer"]
    rows: list[CustomerContactRow] = []
    unnamed_counter = 0
    for row_idx in range(2, sheet.max_row + 1):
        contact_id = parse_int(sheet.cell(row_idx, 1).value)
        contact_code = normalize_text(sheet.cell(row_idx, 2).value)
        contact_name = normalize_text(sheet.cell(row_idx, 3).value)
        organization_name = normalize_text(sheet.cell(row_idx, 4).value)
        address = normalize_text(sheet.cell(row_idx, 5).value)
        phone = normalize_text(sheet.cell(row_idx, 6).value)
        email = normalize_text(sheet.cell(row_idx, 7).value)
        tax_id = normalize_text(sheet.cell(row_idx, 8).value)
        contact_type = normalize_text(sheet.cell(row_idx, 9).value)
        source_channel = normalize_text(sheet.cell(row_idx, 10).value)
        shipping_address = normalize_text(sheet.cell(row_idx, 11).value)
        shipping_email = normalize_text(sheet.cell(row_idx, 12).value)
        shipping_phone = normalize_text(sheet.cell(row_idx, 13).value)
        note_text = normalize_text(sheet.cell(row_idx, 14).value)

        if all(
            item is None
            for item in (
                contact_id,
                contact_name,
                organization_name,
                address,
                phone,
                email,
                tax_id,
                contact_type,
                source_channel,
                shipping_address,
                shipping_email,
                shipping_phone,
                note_text,
            )
        ):
            continue

        if not organization_name and not contact_name:
            unnamed_counter += 1
            fallback = f"Imported Customer {row_idx}-{unnamed_counter}"
        else:
            fallback = f"Imported Customer {row_idx}"

        notes: list[str] = []
        if contact_code:
            notes.append(f"legacy_customer_code: {contact_code}")
        if note_text:
            notes.append(note_text)

        rows.append(
            CustomerContactRow(
                source_contact_id=contact_id,
                name=pick_name(contact_name, organization_name, fallback),
                contact_name=contact_name,
                organization_name=organization_name,
                address=address,
                phone=phone,
                email=email,
                tax_id=tax_id,
                contact_type=contact_type,
                source_channel=source_channel,
                shipping_address=shipping_address,
                shipping_email=shipping_email,
                shipping_phone=shipping_phone,
                notes="\n".join(notes) if notes else None,
            )
        )

    return rows


def parse_suppliers_sheet(workbook: openpyxl.Workbook) -> list[SupplierContactRow]:
    if "Supplier" not in workbook.sheetnames:
        return []

    sheet = workbook["Supplier"]
    rows: list[SupplierContactRow] = []
    unnamed_counter = 0
    for row_idx in range(2, sheet.max_row + 1):
        contact_id = parse_int(sheet.cell(row_idx, 1).value)
        contact_name = normalize_text(sheet.cell(row_idx, 2).value)
        organization_name = normalize_text(sheet.cell(row_idx, 3).value)
        branch_name = normalize_text(sheet.cell(row_idx, 4).value)
        address = normalize_text(sheet.cell(row_idx, 5).value)
        phone = normalize_text(sheet.cell(row_idx, 6).value)
        email = normalize_text(sheet.cell(row_idx, 7).value)
        tax_id = normalize_text(sheet.cell(row_idx, 8).value)
        source_channel = normalize_text(sheet.cell(row_idx, 9).value)
        shipping_address = normalize_text(sheet.cell(row_idx, 10).value)
        shipping_email = normalize_text(sheet.cell(row_idx, 11).value)
        shipping_phone = normalize_text(sheet.cell(row_idx, 12).value)
        note_text = normalize_text(sheet.cell(row_idx, 13).value)

        if all(
            item is None
            for item in (
                contact_id,
                contact_name,
                organization_name,
                branch_name,
                address,
                phone,
                email,
                tax_id,
                source_channel,
                shipping_address,
                shipping_email,
                shipping_phone,
                note_text,
            )
        ):
            continue

        if not organization_name and not contact_name:
            unnamed_counter += 1
            fallback = f"Imported Supplier {row_idx}-{unnamed_counter}"
        else:
            fallback = f"Imported Supplier {row_idx}"

        rows.append(
            SupplierContactRow(
                source_contact_id=contact_id,
                name=pick_name(contact_name, organization_name, fallback),
                contact_name=contact_name,
                organization_name=organization_name,
                branch_name=branch_name,
                address=address,
                phone=phone,
                email=email,
                tax_id=tax_id,
                source_channel=source_channel,
                shipping_address=shipping_address,
                shipping_email=shipping_email,
                shipping_phone=shipping_phone,
                notes=note_text,
            )
        )

    return rows


def sql_literal(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        text = f"{value:.6f}".rstrip("0").rstrip(".")
        return text if text else "0"
    if isinstance(value, (dt.date, dt.datetime)):
        return f"'{value.isoformat()}'"
    escaped = str(value).replace("'", "''")
    return f"N'{escaped}'"


def build_import_sql(
    customer_rows: list[CustomerContactRow],
    supplier_rows: list[SupplierContactRow],
    truncate_first: bool,
) -> str:
    statements: list[str] = [
        "SET XACT_ABORT ON;",
        "BEGIN TRANSACTION;",
    ]

    if truncate_first:
        statements.extend(
            [
                "DELETE FROM dbo.supplier_purchase_history;",
                f"DELETE FROM dbo.suppliers WHERE source_system = {sql_literal(SOURCE_SYSTEM)};",
                f"DELETE FROM dbo.customers WHERE source_system = {sql_literal(SOURCE_SYSTEM)};",
            ]
        )

    for row in customer_rows:
        if row.source_contact_id is not None:
            statements.append(
                (
                    "MERGE dbo.customers AS target "
                    f"USING (SELECT {sql_literal(SOURCE_SYSTEM)} AS source_system, {sql_literal(row.source_contact_id)} AS source_contact_id) AS src "
                    "ON target.source_system = src.source_system AND target.source_contact_id = src.source_contact_id "
                    "WHEN MATCHED THEN UPDATE SET "
                    f"name = {sql_literal(row.name)}, "
                    f"nickname = {sql_literal(row.contact_name)}, "
                    f"organization_name = {sql_literal(row.organization_name)}, "
                    f"address = {sql_literal(row.address)}, "
                    f"phone = {sql_literal(row.phone)}, "
                    f"email = {sql_literal(row.email)}, "
                    f"tax_id = {sql_literal(row.tax_id)}, "
                    f"contact_type = {sql_literal(row.contact_type)}, "
                    f"source_channel = {sql_literal(row.source_channel)}, "
                    f"shipping_address = {sql_literal(row.shipping_address)}, "
                    f"shipping_email = {sql_literal(row.shipping_email)}, "
                    f"shipping_phone = {sql_literal(row.shipping_phone)}, "
                    f"notes = {sql_literal(row.notes)}, "
                    "updated_at_utc = SYSUTCDATETIME() "
                    "WHEN NOT MATCHED THEN INSERT ("
                    "source_system, source_contact_id, name, nickname, organization_name, address, phone, email, tax_id, "
                    "contact_type, source_channel, shipping_address, shipping_email, shipping_phone, notes, customer_since, "
                    "photo_url, photos_json, created_at_utc, updated_at_utc"
                    ") VALUES ("
                    f"{sql_literal(SOURCE_SYSTEM)}, {sql_literal(row.source_contact_id)}, {sql_literal(row.name)}, {sql_literal(row.contact_name)}, "
                    f"{sql_literal(row.organization_name)}, {sql_literal(row.address)}, {sql_literal(row.phone)}, {sql_literal(row.email)}, "
                    f"{sql_literal(row.tax_id)}, {sql_literal(row.contact_type)}, {sql_literal(row.source_channel)}, "
                    f"{sql_literal(row.shipping_address)}, {sql_literal(row.shipping_email)}, {sql_literal(row.shipping_phone)}, "
                    f"{sql_literal(row.notes)}, NULL, NULL, NULL, SYSUTCDATETIME(), SYSUTCDATETIME());"
                )
            )
        else:
            statements.append(
                (
                    "INSERT INTO dbo.customers ("
                    "source_system, source_contact_id, name, nickname, organization_name, address, phone, email, tax_id, "
                    "contact_type, source_channel, shipping_address, shipping_email, shipping_phone, notes, customer_since, "
                    "photo_url, photos_json, created_at_utc, updated_at_utc"
                    ") VALUES ("
                    f"{sql_literal(SOURCE_SYSTEM)}, NULL, {sql_literal(row.name)}, {sql_literal(row.contact_name)}, "
                    f"{sql_literal(row.organization_name)}, {sql_literal(row.address)}, {sql_literal(row.phone)}, {sql_literal(row.email)}, "
                    f"{sql_literal(row.tax_id)}, {sql_literal(row.contact_type)}, {sql_literal(row.source_channel)}, "
                    f"{sql_literal(row.shipping_address)}, {sql_literal(row.shipping_email)}, {sql_literal(row.shipping_phone)}, "
                    f"{sql_literal(row.notes)}, NULL, NULL, NULL, SYSUTCDATETIME(), SYSUTCDATETIME());"
                )
            )

    for row in supplier_rows:
        if row.source_contact_id is not None:
            statements.append(
                (
                    "MERGE dbo.suppliers AS target "
                    f"USING (SELECT {sql_literal(SOURCE_SYSTEM)} AS source_system, {sql_literal(row.source_contact_id)} AS source_contact_id) AS src "
                    "ON target.source_system = src.source_system AND target.source_contact_id = src.source_contact_id "
                    "WHEN MATCHED THEN UPDATE SET "
                    f"name = {sql_literal(row.name)}, "
                    f"contact_name = {sql_literal(row.contact_name)}, "
                    f"organization_name = {sql_literal(row.organization_name)}, "
                    f"branch_name = {sql_literal(row.branch_name)}, "
                    f"address = {sql_literal(row.address)}, "
                    f"phone = {sql_literal(row.phone)}, "
                    f"email = {sql_literal(row.email)}, "
                    f"tax_id = {sql_literal(row.tax_id)}, "
                    f"source_channel = {sql_literal(row.source_channel)}, "
                    f"shipping_address = {sql_literal(row.shipping_address)}, "
                    f"shipping_email = {sql_literal(row.shipping_email)}, "
                    f"shipping_phone = {sql_literal(row.shipping_phone)}, "
                    f"notes = {sql_literal(row.notes)}, "
                    "updated_at_utc = SYSUTCDATETIME() "
                    "WHEN NOT MATCHED THEN INSERT ("
                    "source_system, source_contact_id, name, contact_name, organization_name, branch_name, address, phone, email, tax_id, "
                    "source_channel, shipping_address, shipping_email, shipping_phone, notes, created_at_utc, updated_at_utc"
                    ") VALUES ("
                    f"{sql_literal(SOURCE_SYSTEM)}, {sql_literal(row.source_contact_id)}, {sql_literal(row.name)}, "
                    f"{sql_literal(row.contact_name)}, {sql_literal(row.organization_name)}, {sql_literal(row.branch_name)}, "
                    f"{sql_literal(row.address)}, {sql_literal(row.phone)}, {sql_literal(row.email)}, {sql_literal(row.tax_id)}, "
                    f"{sql_literal(row.source_channel)}, {sql_literal(row.shipping_address)}, {sql_literal(row.shipping_email)}, "
                    f"{sql_literal(row.shipping_phone)}, {sql_literal(row.notes)}, SYSUTCDATETIME(), SYSUTCDATETIME());"
                )
            )
        else:
            statements.append(
                (
                    "INSERT INTO dbo.suppliers ("
                    "source_system, source_contact_id, name, contact_name, organization_name, branch_name, address, phone, email, tax_id, "
                    "source_channel, shipping_address, shipping_email, shipping_phone, notes, created_at_utc, updated_at_utc"
                    ") VALUES ("
                    f"{sql_literal(SOURCE_SYSTEM)}, NULL, {sql_literal(row.name)}, {sql_literal(row.contact_name)}, {sql_literal(row.organization_name)}, "
                    f"{sql_literal(row.branch_name)}, {sql_literal(row.address)}, {sql_literal(row.phone)}, {sql_literal(row.email)}, "
                    f"{sql_literal(row.tax_id)}, {sql_literal(row.source_channel)}, {sql_literal(row.shipping_address)}, "
                    f"{sql_literal(row.shipping_email)}, {sql_literal(row.shipping_phone)}, {sql_literal(row.notes)}, SYSUTCDATETIME(), SYSUTCDATETIME());"
                )
            )

    statements.append("COMMIT TRANSACTION;")
    return "\n".join(statements) + "\n"


def _run_sqlrunner_file(sql_path: Path) -> None:
    root = Path(__file__).resolve().parents[2]
    command = [
        "dotnet",
        "run",
        "--project",
        str(root / "tools" / "db" / "SqlRunner"),
        "-f",
        "net10.0",
        "--",
        "--file",
        str(sql_path),
    ]
    completed = subprocess.run(command, cwd=root, text=True, capture_output=True)
    if completed.stdout.strip():
        print(completed.stdout.strip())
    if completed.returncode != 0:
        if completed.stderr.strip():
            print(completed.stderr.strip())
        raise RuntimeError(f"SqlRunner failed with exit code {completed.returncode}. SQL file: {sql_path}")


def execute_with_sqlrunner(sql_text: str, explicit_sql_path: str | None = None, chunk_size: int = 120) -> Path:
    if explicit_sql_path:
        sql_path = Path(explicit_sql_path).expanduser().resolve()
        sql_path.write_text(sql_text, encoding="utf-8")
    else:
        fd, tmp_name = tempfile.mkstemp(prefix="contacts-import-", suffix=".sql")
        os.close(fd)
        sql_path = Path(tmp_name)
        sql_path.write_text(sql_text, encoding="utf-8")

    lines = [line.strip() for line in sql_text.splitlines() if line.strip()]
    core_lines = [
        line
        for line in lines
        if line not in {"SET XACT_ABORT ON;", "BEGIN TRANSACTION;", "COMMIT TRANSACTION;"}
    ]
    setup_lines = [
        line
        for line in core_lines
        if line.startswith("DELETE FROM dbo.supplier_purchase_history")
        or line.startswith("DELETE FROM dbo.suppliers")
        or line.startswith("DELETE FROM dbo.customers")
    ]
    data_lines = [line for line in core_lines if line not in setup_lines]

    if setup_lines:
        setup_sql = "SET XACT_ABORT ON;\nBEGIN TRANSACTION;\n" + "\n".join(setup_lines) + "\nCOMMIT TRANSACTION;\n"
        fd, setup_name = tempfile.mkstemp(prefix="contacts-import-setup-", suffix=".sql")
        os.close(fd)
        setup_path = Path(setup_name)
        setup_path.write_text(setup_sql, encoding="utf-8")
        _run_sqlrunner_file(setup_path)

    for chunk_index in range(0, len(data_lines), chunk_size):
        chunk_lines = data_lines[chunk_index:chunk_index + chunk_size]
        chunk_sql = "SET XACT_ABORT ON;\nBEGIN TRANSACTION;\n" + "\n".join(chunk_lines) + "\nCOMMIT TRANSACTION;\n"
        fd, chunk_name = tempfile.mkstemp(prefix="contacts-import-chunk-", suffix=".sql")
        os.close(fd)
        chunk_path = Path(chunk_name)
        chunk_path.write_text(chunk_sql, encoding="utf-8")
        _run_sqlrunner_file(chunk_path)

    return sql_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Import customer and supplier workbook data into Azure SQL.")
    parser.add_argument(
        "--excel-path",
        default="/Users/suzieleedhirakul/Downloads/Customers & Suppliers.xlsx",
        help="Path to workbook (.xlsx)",
    )
    parser.add_argument(
        "--no-truncate",
        action="store_true",
        help="Upsert/append workbook rows while keeping existing workbook-sourced records.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse workbook and print row counts without writing to SQL.",
    )
    parser.add_argument(
        "--sql-output",
        help="Optional output path for generated SQL import script.",
    )
    parser.add_argument(
        "--skip-execute",
        action="store_true",
        help="Only generate SQL file without executing it via SqlRunner.",
    )
    args = parser.parse_args()

    workbook_path = Path(args.excel_path).expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    customer_rows = parse_customers_sheet(workbook)
    supplier_rows = parse_suppliers_sheet(workbook)

    print(f"[import] workbook: {workbook_path}")
    print(f"[import] customer rows parsed: {len(customer_rows)}")
    print(f"[import] supplier rows parsed: {len(supplier_rows)}")

    if args.dry_run:
        print("[import] dry-run mode, skipping SQL writes.")
        return 0

    sql_text = build_import_sql(
        customer_rows=customer_rows,
        supplier_rows=supplier_rows,
        truncate_first=not args.no_truncate,
    )

    if args.skip_execute:
        if not args.sql_output:
            raise ValueError("--skip-execute requires --sql-output so the SQL file is persisted.")
        sql_path = Path(args.sql_output).expanduser().resolve()
        sql_path.write_text(sql_text, encoding="utf-8")
        print(f"[import] SQL script generated at: {sql_path}")
    else:
        sql_path = execute_with_sqlrunner(sql_text, explicit_sql_path=args.sql_output)
        print(f"[import] SQL executed from: {sql_path}")

    print(f"[import] imported customer rows: {len(customer_rows)}")
    print(f"[import] imported supplier rows: {len(supplier_rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
