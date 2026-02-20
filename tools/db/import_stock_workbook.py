#!/usr/bin/env python3
"""Import ROJANATORN workbook inventory and usage data into Azure SQL."""

from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Any

import openpyxl

USAGE_SHEET_CATEGORY = {
    "Earrings": "earrings",
    "Neckelet": "necklace",
    "Bracelet": "bracelet",
    "Brooch": "brooch",
    "Clips+Cuffinks": "clips_cufflinks",
    "Ring": "ring",
}

NUMBER_PATTERN = re.compile(r"-?\d+(?:,\d{3})*(?:\.\d+)?")
CT_PATTERN = re.compile(r"(-?\d+(?:\.\d+)?)\s*ct", re.IGNORECASE)
PCS_PATTERN = re.compile(r"(-?\d+(?:\.\d+)?)\s*(?:pcs?|pieces?)", re.IGNORECASE)
SLASH_NUMBER_PATTERN = re.compile(r"/\s*(-?\d+(?:\.\d+)?)")


@dataclass
class InventoryRow:
    source_sheet: str
    source_row: int
    gemstone_number: int | None
    gemstone_number_text: str | None
    gemstone_type: str | None
    weight_pcs_raw: str | None
    shape: str | None
    price_per_ct_raw: str | None
    price_per_piece_raw: str | None
    buying_date: dt.date | None
    buying_date_raw: str | None
    balance_pcs: float | None
    balance_ct: float | None
    use_date: dt.date | None
    use_date_raw: str | None
    owner_name: str | None
    parsed_weight_ct: float | None
    parsed_quantity_pcs: float | None
    parsed_price_per_ct: float | None
    parsed_price_per_piece: float | None


@dataclass
class UsageBatch:
    local_id: int
    source_sheet: str
    source_row: int
    product_category: str
    transaction_date: dt.date | None
    transaction_date_raw: str | None
    requester_name: str | None
    product_code: str | None
    total_amount: float | None


@dataclass
class UsageLine:
    local_batch_id: int
    source_row: int
    gemstone_number: int | None
    gemstone_name: str | None
    used_pcs: float | None
    used_weight_ct: float | None
    unit_price_raw: str | None
    line_amount: float | None
    balance_pcs_after: float | None
    balance_ct_after: float | None
    requester_name: str | None


def normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if text == "" or text == "-":
        return None
    return text


def parse_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if text == "" or text in {"-", "--", "customer"}:
        return None

    text = text.replace("..", ".")
    match = NUMBER_PATTERN.search(text)
    if not match:
        return None

    try:
        return float(match.group(0).replace(",", ""))
    except ValueError:
        return None


def parse_int(value: object) -> int | None:
    parsed = parse_float(value)
    if parsed is None:
        return None
    rounded = int(round(parsed))
    if abs(parsed - rounded) > 1e-6:
        return None
    return rounded


def parse_date(value: object) -> tuple[dt.date | None, str | None]:
    if value is None:
        return None, None

    if isinstance(value, dt.datetime):
        return value.date(), value.date().isoformat()
    if isinstance(value, dt.date):
        return value, value.isoformat()

    text = str(value).strip()
    if text == "" or text in {"-", "--", "#VALUE!"}:
        return None, text if text else None

    formats = (
        "%d/%m/%y",
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%m/%d/%Y",
        "%m/%d/%y",
    )
    for fmt in formats:
        try:
            parsed = dt.datetime.strptime(text, fmt).date()
            return parsed, text
        except ValueError:
            continue

    return None, text


def parse_weight_and_pcs(raw: str | None) -> tuple[float | None, float | None]:
    if not raw:
        return None, None

    lower = raw.lower()
    ct_value: float | None = None
    pcs_value: float | None = None

    ct_match = CT_PATTERN.search(lower)
    if ct_match:
        ct_value = parse_float(ct_match.group(1))

    pcs_match = PCS_PATTERN.search(lower)
    if pcs_match:
        pcs_value = parse_float(pcs_match.group(1))

    slash_match = SLASH_NUMBER_PATTERN.search(lower)
    if slash_match and pcs_value is None:
        pcs_value = parse_float(slash_match.group(1))

    if ct_value is None and pcs_value is None:
        # In some rows the value is just "200.-/pc" or "1pc."
        if "pc" in lower:
            pcs_value = parse_float(lower)
        else:
            ct_value = parse_float(lower)

    return ct_value, pcs_value


def parse_inventory_sheet(workbook: openpyxl.Workbook) -> list[InventoryRow]:
    sheet_name = workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    rows: list[InventoryRow] = []

    for row_idx in range(3, sheet.max_row + 1):
        number_value = sheet.cell(row_idx, 1).value
        gemstone_type = normalize_text(sheet.cell(row_idx, 2).value)
        weight_raw = normalize_text(sheet.cell(row_idx, 3).value)
        shape = normalize_text(sheet.cell(row_idx, 4).value)
        price_ct_raw = normalize_text(sheet.cell(row_idx, 5).value)
        price_piece_raw = normalize_text(sheet.cell(row_idx, 6).value)
        buying_raw_value = sheet.cell(row_idx, 7).value
        balance_pcs = parse_float(sheet.cell(row_idx, 8).value)
        balance_ct = parse_float(sheet.cell(row_idx, 9).value)
        use_raw_value = sheet.cell(row_idx, 10).value
        owner_name = normalize_text(sheet.cell(row_idx, 11).value)

        if (
            number_value is None
            and gemstone_type is None
            and weight_raw is None
            and shape is None
            and price_ct_raw is None
            and price_piece_raw is None
            and buying_raw_value is None
            and balance_pcs is None
            and balance_ct is None
            and use_raw_value is None
            and owner_name is None
        ):
            continue

        gemstone_number = parse_int(number_value)
        gemstone_number_text = normalize_text(number_value)
        buying_date, buying_date_raw = parse_date(buying_raw_value)
        use_date, use_date_raw = parse_date(use_raw_value)
        parsed_weight_ct, parsed_quantity_pcs = parse_weight_and_pcs(weight_raw)

        row = InventoryRow(
            source_sheet=sheet_name,
            source_row=row_idx,
            gemstone_number=gemstone_number,
            gemstone_number_text=gemstone_number_text,
            gemstone_type=gemstone_type,
            weight_pcs_raw=weight_raw,
            shape=shape,
            price_per_ct_raw=price_ct_raw,
            price_per_piece_raw=price_piece_raw,
            buying_date=buying_date,
            buying_date_raw=buying_date_raw,
            balance_pcs=balance_pcs,
            balance_ct=balance_ct,
            use_date=use_date,
            use_date_raw=use_date_raw,
            owner_name=owner_name,
            parsed_weight_ct=parsed_weight_ct,
            parsed_quantity_pcs=parsed_quantity_pcs,
            parsed_price_per_ct=parse_float(price_ct_raw),
            parsed_price_per_piece=parse_float(price_piece_raw),
        )
        rows.append(row)

    return rows


def parse_usage_sheet(workbook: openpyxl.Workbook, sheet_name: str, category: str) -> tuple[list[UsageBatch], list[UsageLine]]:
    sheet = workbook[sheet_name]
    batches: list[UsageBatch] = []
    lines: list[UsageLine] = []
    next_local_id = 1
    current: UsageBatch | None = None

    def flush_current() -> None:
        nonlocal current
        if current is not None:
            batches.append(current)
            current = None

    for row_idx in range(3, sheet.max_row + 1):
        date_value = sheet.cell(row_idx, 1).value
        requester_name = normalize_text(sheet.cell(row_idx, 2).value)
        product_code = normalize_text(sheet.cell(row_idx, 3).value)
        gemstone_number = parse_int(sheet.cell(row_idx, 4).value)
        gemstone_name = normalize_text(sheet.cell(row_idx, 5).value)
        used_pcs = parse_float(sheet.cell(row_idx, 6).value)
        used_weight_ct = parse_float(sheet.cell(row_idx, 7).value)
        unit_price_raw = normalize_text(sheet.cell(row_idx, 8).value)
        line_amount = parse_float(sheet.cell(row_idx, 9).value)
        total_amount = parse_float(sheet.cell(row_idx, 10).value)
        balance_pcs_after = parse_float(sheet.cell(row_idx, 11).value)
        balance_ct_after = parse_float(sheet.cell(row_idx, 12).value)
        transaction_date, transaction_date_raw = parse_date(date_value)

        # Skip header-like rows that appear below row 1.
        if requester_name in {"ลูกค้า", "customer"} and product_code is None and gemstone_name is None:
            continue
        if gemstone_name in {"ชื่อพลอย"}:
            continue

        has_line = any(
            value is not None
            for value in (gemstone_number, gemstone_name, used_pcs, used_weight_ct, unit_price_raw, line_amount)
        )

        has_batch_marker = product_code is not None or total_amount is not None or transaction_date is not None
        if not has_line and not has_batch_marker:
            continue

        should_start_new_batch = product_code is not None
        if should_start_new_batch:
            flush_current()

        if current is None and (has_line or has_batch_marker):
            current = UsageBatch(
                local_id=next_local_id,
                source_sheet=sheet_name,
                source_row=row_idx,
                product_category=category,
                transaction_date=transaction_date,
                transaction_date_raw=transaction_date_raw,
                requester_name=requester_name,
                product_code=product_code,
                total_amount=total_amount,
            )
            next_local_id += 1
        elif current is not None:
            if current.transaction_date is None and transaction_date is not None:
                current.transaction_date = transaction_date
                current.transaction_date_raw = transaction_date_raw
            if current.requester_name is None and requester_name is not None:
                current.requester_name = requester_name
            if current.product_code is None and product_code is not None:
                current.product_code = product_code
            if total_amount is not None:
                current.total_amount = total_amount

        if current is None:
            continue

        if has_line:
            lines.append(
                UsageLine(
                    local_batch_id=current.local_id,
                    source_row=row_idx,
                    gemstone_number=gemstone_number,
                    gemstone_name=gemstone_name,
                    used_pcs=used_pcs,
                    used_weight_ct=used_weight_ct,
                    unit_price_raw=unit_price_raw,
                    line_amount=line_amount,
                    balance_pcs_after=balance_pcs_after,
                    balance_ct_after=balance_ct_after,
                    requester_name=requester_name,
                )
            )

    flush_current()

    # Keep only batches that have usable identifiers or lines.
    line_batch_ids = {line.local_batch_id for line in lines}
    filtered_batches = [
        batch
        for batch in batches
        if batch.local_id in line_batch_ids or batch.product_code is not None or batch.total_amount is not None
    ]
    filtered_local_ids = {batch.local_id for batch in filtered_batches}
    filtered_lines = [line for line in lines if line.local_batch_id in filtered_local_ids]
    return filtered_batches, filtered_lines


def sql_literal(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value != value:  # NaN
            return "NULL"
        text = f"{value:.6f}".rstrip("0").rstrip(".")
        return text if text else "0"
    if isinstance(value, (dt.date, dt.datetime)):
        date_value = value.date() if isinstance(value, dt.datetime) else value
        return f"'{date_value.isoformat()}'"

    escaped = str(value).replace("'", "''")
    return f"N'{escaped}'"


def build_import_sql(
    inventory_rows: Iterable[InventoryRow],
    usage_batches: Iterable[UsageBatch],
    usage_lines: Iterable[UsageLine],
    truncate_first: bool,
) -> str:
    inventory_rows = list(inventory_rows)
    usage_batches = list(usage_batches)
    usage_lines = list(usage_lines)

    statements: list[str] = [
        "SET XACT_ABORT ON;",
        "BEGIN TRANSACTION;",
    ]

    if truncate_first:
        statements.extend(
            [
                "DELETE FROM dbo.gem_usage_lines;",
                "DELETE FROM dbo.gem_usage_batches;",
                "DELETE FROM dbo.gem_inventory_items;",
            ]
        )

    for row in inventory_rows:
        statements.append(
            (
                "INSERT INTO dbo.gem_inventory_items ("
                "source_sheet, source_row, gemstone_number, gemstone_number_text, gemstone_type, weight_pcs_raw, "
                "shape, price_per_ct_raw, price_per_piece_raw, buying_date, buying_date_raw, balance_pcs, balance_ct, "
                "use_date, use_date_raw, owner_name, parsed_weight_ct, parsed_quantity_pcs, parsed_price_per_ct, "
                "parsed_price_per_piece"
                ") VALUES ("
                f"{sql_literal(row.source_sheet)}, "
                f"{sql_literal(row.source_row)}, "
                f"{sql_literal(row.gemstone_number)}, "
                f"{sql_literal(row.gemstone_number_text)}, "
                f"{sql_literal(row.gemstone_type)}, "
                f"{sql_literal(row.weight_pcs_raw)}, "
                f"{sql_literal(row.shape)}, "
                f"{sql_literal(row.price_per_ct_raw)}, "
                f"{sql_literal(row.price_per_piece_raw)}, "
                f"{sql_literal(row.buying_date)}, "
                f"{sql_literal(row.buying_date_raw)}, "
                f"{sql_literal(row.balance_pcs)}, "
                f"{sql_literal(row.balance_ct)}, "
                f"{sql_literal(row.use_date)}, "
                f"{sql_literal(row.use_date_raw)}, "
                f"{sql_literal(row.owner_name)}, "
                f"{sql_literal(row.parsed_weight_ct)}, "
                f"{sql_literal(row.parsed_quantity_pcs)}, "
                f"{sql_literal(row.parsed_price_per_ct)}, "
                f"{sql_literal(row.parsed_price_per_piece)}"
                ");"
            )
        )

    usage_batches = list(usage_batches)
    usage_lines = list(usage_lines)
    if truncate_first:
        batch_id_map = {batch.local_id: batch.local_id for batch in usage_batches}
    else:
        seed = 1_000_000 + (int(dt.datetime.now(dt.timezone.utc).timestamp()) % 1_000_000) * 1000
        batch_id_map = {batch.local_id: seed + index for index, batch in enumerate(usage_batches)}

    if usage_batches:
        statements.append("SET IDENTITY_INSERT dbo.gem_usage_batches ON;")
        for batch in usage_batches:
            explicit_id = batch_id_map[batch.local_id]
            statements.append(
                (
                    "INSERT INTO dbo.gem_usage_batches ("
                    "id, source_sheet, source_row, product_category, transaction_date, transaction_date_raw, "
                    "requester_name, product_code, total_amount"
                    ") VALUES ("
                    f"{sql_literal(explicit_id)}, "
                    f"{sql_literal(batch.source_sheet)}, "
                    f"{sql_literal(batch.source_row)}, "
                    f"{sql_literal(batch.product_category)}, "
                    f"{sql_literal(batch.transaction_date)}, "
                    f"{sql_literal(batch.transaction_date_raw)}, "
                    f"{sql_literal(batch.requester_name)}, "
                    f"{sql_literal(batch.product_code)}, "
                    f"{sql_literal(batch.total_amount)}"
                    ");"
                )
            )
        statements.append("SET IDENTITY_INSERT dbo.gem_usage_batches OFF;")

    for line in usage_lines:
        batch_id = batch_id_map.get(line.local_batch_id)
        if batch_id is None:
            continue
        statements.append(
            (
                "INSERT INTO dbo.gem_usage_lines ("
                "batch_id, source_row, gemstone_number, gemstone_name, used_pcs, used_weight_ct, unit_price_raw, "
                "line_amount, balance_pcs_after, balance_ct_after, requester_name"
                ") VALUES ("
                f"{sql_literal(batch_id)}, "
                f"{sql_literal(line.source_row)}, "
                f"{sql_literal(line.gemstone_number)}, "
                f"{sql_literal(line.gemstone_name)}, "
                f"{sql_literal(line.used_pcs)}, "
                f"{sql_literal(line.used_weight_ct)}, "
                f"{sql_literal(line.unit_price_raw)}, "
                f"{sql_literal(line.line_amount)}, "
                f"{sql_literal(line.balance_pcs_after)}, "
                f"{sql_literal(line.balance_ct_after)}, "
                f"{sql_literal(line.requester_name)}"
                ");"
            )
        )

    statements.append("COMMIT TRANSACTION;")
    return "\n".join(statements) + "\n"


def execute_with_sqlrunner(sql_text: str, explicit_sql_path: str | None = None) -> Path:
    root = Path(__file__).resolve().parents[2]
    if explicit_sql_path:
        sql_path = Path(explicit_sql_path).expanduser().resolve()
        sql_path.write_text(sql_text, encoding="utf-8")
    else:
        fd, tmp_name = tempfile.mkstemp(prefix="stock-import-", suffix=".sql")
        os.close(fd)
        sql_path = Path(tmp_name)
        sql_path.write_text(sql_text, encoding="utf-8")

    env = os.environ.copy()
    env["PATH"] = f"/opt/homebrew/opt/dotnet@8/bin:{env.get('PATH', '')}"
    env["DOTNET_ROOT"] = "/opt/homebrew/opt/dotnet@8/libexec"

    command = [
        "dotnet",
        "run",
        "--project",
        str(root / "tools" / "db" / "SqlRunner"),
        "--",
        "--file",
        str(sql_path),
    ]
    completed = subprocess.run(command, cwd=root, env=env, text=True, capture_output=True)
    if completed.stdout.strip():
        print(completed.stdout.strip())
    if completed.returncode != 0:
        if completed.stderr.strip():
            print(completed.stderr.strip())
        raise RuntimeError(f"SqlRunner failed with exit code {completed.returncode}. SQL file: {sql_path}")
    return sql_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Import gemstone workbook data into Azure SQL.")
    parser.add_argument(
        "--excel-path",
        default="/Users/suzieleedhirakul/Downloads/ROJANATORN GEMS STOCK 2026.xlsx",
        help="Path to workbook (.xlsx)",
    )
    parser.add_argument(
        "--no-truncate",
        action="store_true",
        help="Append into target tables instead of truncating first.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse workbook and print row counts without writing to SQL.",
    )
    parser.add_argument(
        "--sql-output",
        help="Optional output path for generated SQL import script.",
    )
    parser.add_argument(
        "--skip-execute",
        action="store_true",
        help="Only generate SQL file without executing it via SqlRunner.",
    )
    args = parser.parse_args()

    workbook_path = Path(args.excel_path).expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    inventory_rows = parse_inventory_sheet(workbook)

    usage_batches: list[UsageBatch] = []
    usage_lines: list[UsageLine] = []
    for sheet_name, category in USAGE_SHEET_CATEGORY.items():
        if sheet_name not in workbook.sheetnames:
            continue
        batches, lines = parse_usage_sheet(workbook, sheet_name, category)
        local_id_map: dict[int, int] = {}
        for batch in batches:
            next_global_id = len(usage_batches) + 1
            local_id_map[batch.local_id] = next_global_id
            batch.local_id = next_global_id
            usage_batches.append(batch)
        for line in lines:
            remapped = local_id_map.get(line.local_batch_id)
            if remapped is None:
                continue
            line.local_batch_id = remapped
            usage_lines.append(line)

    print(f"[import] workbook: {workbook_path}")
    print(f"[import] inventory rows parsed: {len(inventory_rows)}")
    print(f"[import] usage batches parsed: {len(usage_batches)}")
    print(f"[import] usage lines parsed: {len(usage_lines)}")

    if args.dry_run:
        print("[import] dry-run mode, skipping SQL writes.")
        return 0

    sql_text = build_import_sql(
        inventory_rows=inventory_rows,
        usage_batches=usage_batches,
        usage_lines=usage_lines,
        truncate_first=not args.no_truncate,
    )

    if args.skip_execute:
        if not args.sql_output:
            raise ValueError("--skip-execute requires --sql-output so the SQL file is persisted.")
        sql_path = Path(args.sql_output).expanduser().resolve()
        sql_path.write_text(sql_text, encoding="utf-8")
        print(f"[import] SQL script generated at: {sql_path}")
    else:
        sql_path = execute_with_sqlrunner(sql_text, explicit_sql_path=args.sql_output)
        print(f"[import] SQL executed from: {sql_path}")

    print(f"[import] inserted inventory rows: {len(inventory_rows)}")
    print(f"[import] inserted usage batches: {len(usage_batches)}")
    print(f"[import] inserted usage lines: {len(usage_lines)}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
